"""
parse_kpi_excel.py
Convierte TABLA KPIs 4G_1h.xlsx a kpi_data.json para el panel de KPIs de la app.

Salida:
{
  "kpis": {
    "GALX0904M1A": {
      "province": "LA CORUNA",
      "node": "GALX0904",
      "hourly": [
        {"date": "21-feb", "hora": "11:00", "cell_avail": 100, "prb_dl": 48.8, ...},
        ...
      ]
    },
    ...
  },
  "kpi_meta": [
    {"key": "cell_avail", "label": "Cell Avail %", "unit": "%", "good_direction": "high", "warn_below": 99, "crit_below": 95},
    ...
  ]
}
"""

import json, sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("pip install openpyxl")

EXCEL_PATH = Path("C:/Users/pparedes/Downloads/TABLA KPIs 4G_1h.xlsx")
OUT_PATH   = Path("C:/Users/pparedes/OneDrive - Kenmei Technologies/Escritorio/Interference tool/topology-explorer/public/kpi_data.json")

# Mapeo columna Excel -> key JSON
COL_MAP = {
    "Cell Avail":                  ("cell_avail",      "%",    "high",  99,   95),
    "VoLTE (minutos)":             ("volte_min",        "min",  "high",  None, None),
    "DL MB Volte":                 ("volte_dl_mb",      "MB",   "high",  None, None),
    "PRB_DL":                      ("prb_dl",           "%",    "low",   70,   90),
    "DL_Aggregates_Throu_User":    ("dl_tput_num",      "",     None,    None, None),
    "DL_Trou_Time_User":           ("dl_tput_den",      "",     None,    None, None),
    "ERAB_EST_SUCC":               ("erab_succ",        "",     None,    None, None),
    "ERAB_EST_ATT":                ("erab_att",         "",     None,    None, None),
    "S1_SR_SUCC":                  ("s1_succ",          "",     None,    None, None),
    "S1_SR_ATT":                   ("s1_att",           "",     None,    None, None),
    "RRC_SUCC":                    ("rrc_succ",         "",     None,    None, None),
    "RRC_ATT":                     ("rrc_att",          "",     None,    None, None),
    "PDCCH_Util %":                ("pdcch_util",       "%",    "low",   70,   85),
    "Pmrrcconnlevsum":             ("rrc_users_sum",    "",     None,    None, None),
    "Pmrrcconnlevsamp":            ("rrc_users_samp",   "",     None,    None, None),
}

# KPIs derivados que calcularemos
KPI_META = [
    {"key": "cell_avail",    "label": "Disponibilidad",         "unit": "%",    "good_direction": "high", "warn_below": 99,   "crit_below": 95},
    {"key": "erab_access",   "label": "ERAB Accesibilidad",     "unit": "%",    "good_direction": "high", "warn_below": 98,   "crit_below": 95},
    {"key": "prb_dl",        "label": "PRB DL Utilización",     "unit": "%",    "good_direction": "low",  "warn_above": 70,   "crit_above": 85},
    {"key": "dl_tput_mbps",  "label": "DL Throughput",          "unit": "Mbps", "good_direction": "high", "warn_below": None, "crit_below": None},
    {"key": "pdcch_util",    "label": "PDCCH Congestión",       "unit": "%",    "good_direction": "low",  "warn_above": 70,   "crit_above": 85},
    {"key": "rrc_users",     "label": "Usuarios RRC promedio",  "unit": "",     "good_direction": None,   "warn_below": None, "crit_below": None},
    {"key": "volte_min",     "label": "VoLTE (minutos)",        "unit": "min",  "good_direction": "high", "warn_below": None, "crit_below": None},
]

def safe_div(n, d):
    if n is None or d is None: return None
    try:
        nd, dd = float(n), float(d)
        return round(nd / dd, 4) if dd != 0 else None
    except:
        return None

def safe_float(v):
    if v is None: return None
    try: return float(v)
    except: return None

def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb["TABLA KPIs 4G_1h(12)"]

    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col_idx = {h: i for i, h in enumerate(headers) if h}

    result = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        node    = str(row[col_idx.get("Nodo", 0)] or "").strip()
        cell_id = str(row[col_idx.get("Celda", 2)] or "").strip()
        dia     = str(row[col_idx.get("Dia", 4)] or "").strip()
        hora    = str(row[col_idx.get("Hora", 5)] or "").strip()
        prov    = str(row[col_idx.get("provincia: Descending", 6)] or "").strip()

        if not cell_id or not hora: continue

        if cell_id not in result:
            result[cell_id] = {"province": prov, "node": node, "hourly": []}

        entry = {"date": dia, "hora": hora}

        # Raw columns
        for col_name, (key, *_) in COL_MAP.items():
            idx = col_idx.get(col_name)
            if idx is not None:
                entry[key] = safe_float(row[idx])

        # Derived KPIs
        entry["erab_access"]  = safe_div(entry.get("erab_succ"), entry.get("erab_att"))
        if entry["erab_access"]: entry["erab_access"] = round(entry["erab_access"] * 100, 2)

        entry["dl_tput_mbps"] = safe_div(entry.get("dl_tput_num"), entry.get("dl_tput_den"))
        if entry["dl_tput_mbps"]: entry["dl_tput_mbps"] = round(entry["dl_tput_mbps"] / 1e6 * 8, 2)

        entry["rrc_users"]    = safe_div(entry.get("rrc_users_sum"), entry.get("rrc_users_samp"))
        if entry["rrc_users"]: entry["rrc_users"] = round(entry["rrc_users"], 1)

        # Clean internal keys not needed in frontend
        for k in ["erab_succ","erab_att","s1_succ","s1_att","rrc_succ","rrc_att","dl_tput_num","dl_tput_den","rrc_users_sum","rrc_users_samp"]:
            entry.pop(k, None)

        result[cell_id]["hourly"].append(entry)

    output = {"kpis": result, "kpi_meta": KPI_META}
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT_PATH, "w", encoding="utf-8") as f:
        json.dump(output, f, separators=(",", ":"), ensure_ascii=False)

    total_rows = sum(len(v["hourly"]) for v in result.values())
    print(f"Celdas: {len(result)}, Filas: {total_rows}, Guardado: {OUT_PATH}")
    print("Celdas:", sorted(result.keys()))

if __name__ == "__main__":
    main()
